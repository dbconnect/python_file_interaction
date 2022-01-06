import openpyxl


def read_excel():
    workbook = openpyxl.load_workbook('persons.xlsx', read_only=True)
    first_sheet = workbook.sheetnames[0]
    sheet = workbook[first_sheet]
    data = []
    for i in range(1, sheet.max_row + 1):
        row = []
        for j in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(i, j).value
            row.append(cell_value)
        data.append(row)
    print(data)


if __name__ == '__main__':
    read_excel()
