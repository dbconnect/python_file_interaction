import openpyxl
from openpyxl.styles import Font


def write_excel():
    data = [
        ['first_name', 'last_name'],
        ['Harvey', 'Specter'],
        ['Michael', 'Ross']
    ]
    workbook = openpyxl.Workbook()
    workbook.create_sheet()
    first_sheet = workbook.sheetnames[0]
    sheet = workbook[first_sheet]
    i = 1
    for row in data:
        j = 1
        for data in row:
            cell = sheet.cell(i, j)
            cell.value = data
            if i == 1:
                cell.font = Font(bold=True)
            j += 1
        i += 1
    workbook.save("output.xlsx")


if __name__ == '__main__':
    write_excel()
